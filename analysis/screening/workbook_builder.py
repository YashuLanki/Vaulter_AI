"""
analysis/screening/workbook_builder.py
-----------------------------------------
Builds ONE combined openpyxl workbook with all 4 screening phases as
separate sheets, in a fixed order:

  1. Phase4_Final_Verification  (matrix: rows=categories, cols=finalists)
  2. Phase3_Deep_Analysis        (matrix: rows=categories, cols=top-N)
  3. Phase2_Ranked                (row-per-listing: Phase 1 survivors)
  4. Phase1_Screening              (row-per-listing: all raw listings)

This sheet order and these exact sheet names matter -- the dashboard
(analysis/screening/dashboard/vaulter_dashboard.html) reads them by name.
"""

import os
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from . import config as screening_config

PHASE4_ROW_LABELS = [
    "Composite_Score", "Phase3_Recommendation", "VISUAL_INSPECTION",
    "GROUND_TRUTH_FINDINGS", "RISK_ASSESSMENT", "FINAL_RECOMMENDATION",
    "REMAINING_DILIGENCE_ITEMS",
]

PHASE3_ROW_LABELS = [
    "Composite_Score", "STRENGTHS_AND_RISKS", "ENTITLEMENT_RISK",
    "RECOMMENDATION", "MOIC_FIT", "RED_FLAGS",
]

PHASE2_DISPLAY_COLS = [
    "Composite_Score", "Screening_Status", "Score_DaysOnMarket", "Score_Price",
    "Score_LandUseCategory", "Score_DevelopedEnv", "Score_FloodRisk",
    "Screening_Reasons",
] + screening_config.OUTPUT_COLUMNS

PHASE1_DISPLAY_COLS = [
    "Screening_Status", "Screening_Reasons", "Flag_Count",
] + screening_config.OUTPUT_COLUMNS


def _style_matrix_sheet(ws: Worksheet, row_labels: list, addresses: list):
    """Applies the same Font/Alignment styling costar's deep_analysis.py /
    final_verification.py used for their matrix sheets."""
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(addresses) + 2):
        col_letter = openpyxl.utils.get_column_letter(j)
        ws.column_dimensions[col_letter].width = 45

    ws.row_dimensions[1].height = 20
    for i in range(2, len(row_labels) + 2):
        ws.row_dimensions[i].height = 150
    ws.freeze_panes = "B2"


def _write_matrix_sheet(wb: openpyxl.Workbook, title: str, row_labels: list, data: dict) -> Worksheet:
    """data: {address: {row_label: value}}"""
    ws = wb.create_sheet(title=title)
    addresses = list(data.keys())

    for j, addr in enumerate(addresses, start=2):
        ws.cell(row=1, column=j, value=addr).font = Font(bold=True)

    for i, label in enumerate(row_labels, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        for j, addr in enumerate(addresses, start=2):
            value = data.get(addr, {}).get(label, "")
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    _style_matrix_sheet(ws, row_labels, addresses)
    return ws


def _write_row_per_listing_sheet(wb: openpyxl.Workbook, title: str, df: pd.DataFrame, display_cols: list) -> Worksheet:
    ws = wb.create_sheet(title=title)
    cols = [c for c in display_cols if c in df.columns]

    for j, col in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=col).font = Font(bold=True)

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(cols, start=1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            if col in ("Screening_Reasons",):
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for j, col in enumerate(cols, start=1):
        col_letter = openpyxl.utils.get_column_letter(j)
        ws.column_dimensions[col_letter].width = 32 if col == "Screening_Reasons" else 18

    ws.freeze_panes = "A2"
    return ws


def build_combined_workbook(
    screening_df: pd.DataFrame,
    ranked_df: pd.DataFrame,
    deep_analyses: dict,
    final_result: dict,
    output_path: Path,
) -> Path:
    """
    Creates ONE openpyxl workbook with exactly 4 sheets, in this order:
      1. Phase4_Final_Verification
      2. Phase3_Deep_Analysis
      3. Phase2_Ranked
      4. Phase1_Screening
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet openpyxl creates -- we add our own 4 in order.
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── Sheet 1: Phase4_Final_Verification ──────────────────────────
    if final_result.get("skipped"):
        finalists = final_result.get("finalists", [])
        phase4_data = {}
        for addr in finalists:
            composite = None
            recommendation = ""
            if addr in deep_analyses:
                composite = deep_analyses[addr].get("Composite_Score")
                recommendation = deep_analyses[addr].get("RECOMMENDATION", "")
            row_values = {
                "Composite_Score": composite,
                "Phase3_Recommendation": recommendation,
            }
            for label in PHASE4_ROW_LABELS:
                if label not in row_values:
                    row_values[label] = "Skipped — no Google Maps API key configured"
            phase4_data[addr] = row_values
    else:
        phase4_data = final_result.get("analyses", {})

    _write_matrix_sheet(wb, "Phase4_Final_Verification", PHASE4_ROW_LABELS, phase4_data)

    # ── Sheet 2: Phase3_Deep_Analysis ────────────────────────────────
    _write_matrix_sheet(wb, "Phase3_Deep_Analysis", PHASE3_ROW_LABELS, deep_analyses)

    # ── Sheet 3: Phase2_Ranked ───────────────────────────────────────
    _write_row_per_listing_sheet(wb, "Phase2_Ranked", ranked_df, PHASE2_DISPLAY_COLS)

    # ── Sheet 4: Phase1_Screening ────────────────────────────────────
    _write_row_per_listing_sheet(wb, "Phase1_Screening", screening_df, PHASE1_DISPLAY_COLS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Save to a temp file in the same directory, then rename into place --
    # Path.replace() is atomic on both POSIX and Windows for a same-
    # filesystem rename, so a reader (a teammate's OneDrive-synced copy,
    # or this machine's own dashboard) can never catch a half-written
    # workbook mid-save (see C5 in docs/MULTI_USER_TRANSITION.md).
    tmp_path = output_path.with_name(f"{output_path.name}.tmp{os.getpid()}")
    wb.save(tmp_path)
    tmp_path.replace(output_path)
    return output_path
