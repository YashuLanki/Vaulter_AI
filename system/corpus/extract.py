"""
corpus/extract.py
-----------------
Turn one document from the firm's library into plain text for Claude.

Moved from the old `ingestion/extractor.py`, which fed a chunker and an
embedder. Nothing chunks or embeds any more -- the text goes straight back to
the conversation -- so two things changed:

  * The Excel path no longer stamps `page_count = 9999`. That sentinel existed
    purely to trigger an 8000-character chunk tier in `config.CHUNK_TIERS` so
    CoStar rows wouldn't be split mid-row. There is no chunker to signal.
  * `.docx` is supported (via mammoth). The library is full of Word documents;
    the old extractor skipped them because only the email-attachment path
    handled Word, and that path is gone.

Reading a file here HYDRATES it -- OneDrive downloads the bytes on first
access. That is the whole reason `corpus.index.search` works on filenames
instead of content: search stays free, and only the specific document someone
chose to open costs a download.

Known gap: `.msg` (archived Outlook messages) is not supported. The library
has a lot of them and they likely carry real correspondence history, but
reading them needs another dependency (`extract-msg`). Flagged, not silently
skipped -- `read_document` says so explicitly when asked for one.
"""

import itertools
import logging
from datetime import datetime
from pathlib import Path

import pdfplumber
import pytesseract
from pdf2image import convert_from_path

from config import TESSERACT_PATH, POPPLER_PATH
from corpus.index import is_online_only, resolve_in_corpus

pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

log = logging.getLogger("vaulter.corpus.extract")

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls", ".csv", ".txt", ".md"}

# Read but explained rather than silently returning nothing.
_KNOWN_UNSUPPORTED = {
    ".msg": "an archived Outlook message (needs the extract-msg package, not installed)",
    ".eml": "an archived email (no reader installed)",
    ".pptx": "a PowerPoint deck (no reader installed)",
    ".dwg": "a CAD drawing",
    ".shp": "a GIS shapefile",
}


def is_supported(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


def read_document(rel_path: str, max_chars: int = 200_000) -> tuple[str, dict]:
    """
    Read one document out of the library and return (text, metadata).

    `rel_path` is corpus-relative and is resolved through the scope guard, so
    this cannot be pointed at anything outside the document library.

    Args:
        rel_path:  path relative to CORPUS_DIR, as returned by corpus.search
        max_chars: truncate beyond this, with a marker. Guards against handing
                   back a 400-page appraisal in one blob.
    """
    path = resolve_in_corpus(rel_path)
    if not path.is_file():
        raise FileNotFoundError(f"No such document in the library: {rel_path}")

    ext = path.suffix.lower()
    metadata = {
        "path": rel_path,
        "filename": path.name,
        "file_type": ext,
        "size_bytes": path.stat().st_size,
        "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        "was_online_only": is_online_only(path),
        "read_at": datetime.now().isoformat(timespec="seconds"),
        "page_count": 0,
        "has_tables": False,
        "ocr_used": False,
        "truncated": False,
    }

    if ext not in SUPPORTED_EXTENSIONS:
        described = _KNOWN_UNSUPPORTED.get(ext, f"an unsupported file type ({ext})")
        return f"[Cannot read {path.name}: it is {described}.]", metadata

    if ext == ".pdf":
        text, metadata = _extract_pdf(path, metadata)
    elif ext == ".docx":
        text, metadata = _extract_docx(path, metadata)
    elif ext in (".xlsx", ".xls"):
        text, metadata = _extract_excel(path, metadata)
    elif ext == ".csv":
        text, metadata = _extract_csv(path, metadata)
    else:
        text, metadata = _extract_txt(path, metadata)

    if len(text) > max_chars:
        metadata["truncated"] = True
        metadata["full_length"] = len(text)
        text = (
            text[:max_chars]
            + f"\n\n[Truncated at {max_chars:,} of {len(text):,} characters. "
              f"Ask for a specific section or page range to see more.]"
        )
    return text, metadata


# ─── PDF ──────────────────────────────────────────────────────────────────────

def _extract_pdf(path: Path, metadata: dict) -> tuple[str, dict]:
    """
    Extract each page with pdfplumber. Any individual page that yields no
    text layer (e.g. a scanned image page mixed into an otherwise digital PDF)
    falls back to Tesseract OCR for that page only, so a mostly-digital PDF
    with a few scanned pages doesn't silently drop those pages -- only a
    whole-document OCR fallback would have caught an all-scanned PDF, missing
    the mixed case.
    """
    full_text = []

    with pdfplumber.open(path) as pdf:
        metadata["page_count"] = len(pdf.pages)

        if pdf.metadata:
            metadata["pdf_title"]  = pdf.metadata.get("Title", "") or ""
            metadata["pdf_author"] = pdf.metadata.get("Author", "") or ""

        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()

            if text and text.strip():
                full_text.append(f"[Page {page_num}]\n{text.strip()}")
            else:
                log.info(f"  Page {page_num} has no text layer — running OCR...")
                # Render only this one page, not the whole document -- a
                # mostly-digital PDF with one scanned/blank page would
                # otherwise pay to rasterize every page at 300 DPI just to
                # OCR the one that needs it.
                page_images = convert_from_path(
                    str(path), dpi=300, poppler_path=POPPLER_PATH,
                    first_page=page_num, last_page=page_num,
                )
                metadata["ocr_used"] = True

                if page_images:
                    ocr_text = pytesseract.image_to_string(page_images[0], lang="eng")
                    if ocr_text.strip():
                        full_text.append(f"[Page {page_num} - OCR]\n{ocr_text.strip()}")

            tables = page.extract_tables()
            if tables:
                metadata["has_tables"] = True
                for table in tables:
                    table_text = _table_to_text(table, page_num)
                    if table_text:
                        full_text.append(table_text)

    return "\n\n".join(full_text), metadata


def _table_to_text(table: list, page_num: int) -> str:
    """Convert a pdfplumber table (list of lists) into readable plain text."""
    if not table:
        return ""
    lines = [f"[Table on Page {page_num}]"]
    for row in table:
        cleaned = [str(cell).strip() if cell else "" for cell in row]
        lines.append(" | ".join(cleaned))
    return "\n".join(lines)


# ─── Word ─────────────────────────────────────────────────────────────────────

def _extract_docx(path: Path, metadata: dict) -> tuple[str, dict]:
    """Extract a Word document as markdown, which keeps headings and tables."""
    import mammoth

    try:
        with open(path, "rb") as f:
            result = mammoth.convert_to_markdown(f)
        metadata["page_count"] = 1
        metadata["has_tables"] = "|" in result.value
        return result.value, metadata
    except Exception as e:
        log.error(f"  [ERROR] Failed to extract Word file: {e}")
        return f"[Could not read {path.name}: {e}]", metadata


# ─── Excel ────────────────────────────────────────────────────────────────────

def _extract_excel(path: Path, metadata: dict) -> tuple[str, dict]:
    """
    Extract all sheets and cells from an Excel file (.xlsx or .xls).
    Each sheet is converted to readable plain text with rows and columns.
    """
    import openpyxl

    full_text = []
    metadata["has_tables"] = True

    try:
        wb_formulas = None
        if path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(path, data_only=True)
            # data_only=True returns None for any formula cell that was never
            # recalculated/saved by Excel (e.g. a workbook generated
            # programmatically and never opened in Excel) -- a row made up
            # entirely of such cells looks completely empty and would be
            # silently skipped below, even though it has real (just uncached)
            # data. Load a second, formula-preserving copy so we can tell
            # "genuinely blank row" apart from "all-uncalculated-formula row"
            # and fall back to showing the formula text itself rather than
            # losing the row entirely.
            wb_formulas = openpyxl.load_workbook(path, data_only=False)
        else:
            import xlrd
            xls_wb = xlrd.open_workbook(str(path))
            wb = _convert_xls_to_openpyxl(xls_wb)

        metadata["page_count"] = len(wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines = [f"[Sheet: {sheet_name}]"]

            formula_rows = wb_formulas[sheet_name].iter_rows(values_only=True) if wb_formulas else iter(())
            for row, formula_row in itertools.zip_longest(ws.iter_rows(values_only=True), formula_rows, fillvalue=()):
                if all(cell is None for cell in row):
                    if any(isinstance(c, str) and c.startswith("=") for c in formula_row):
                        # Not actually empty -- every cell is an uncalculated
                        # formula. Show the formula text since we can't
                        # evaluate it ourselves.
                        cleaned = [str(c) if c is not None else "" for c in formula_row]
                        sheet_lines.append(" | ".join(cleaned))
                    continue
                cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
                sheet_lines.append(" | ".join(cleaned))

            if len(sheet_lines) > 1:
                full_text.append("\n".join(sheet_lines))

        log.info(f"  Extracted {metadata['page_count']} sheet(s) from Excel file")

    except Exception as e:
        log.error(f"  [ERROR] Failed to extract Excel file: {e}")
        return f"[Could not read {path.name}: {e}]", metadata

    return "\n\n".join(full_text), metadata


def _convert_xls_to_openpyxl(xls_wb):
    """Convert an xlrd workbook to openpyxl format for uniform processing."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_idx in range(xls_wb.nsheets):
        xls_sheet = xls_wb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=xls_sheet.name)
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                ws.cell(row=row + 1, column=col + 1, value=xls_sheet.cell_value(row, col))

    return wb


# ─── CSV / plain text ─────────────────────────────────────────────────────────

def _extract_csv(path: Path, metadata: dict) -> tuple[str, dict]:
    """Extract a CSV as readable plain text via pandas."""
    import pandas as pd

    metadata["has_tables"] = True

    try:
        try:
            df = pd.read_csv(path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, encoding="latin-1")

        metadata["page_count"] = 1

        lines = [
            f"[CSV File: {path.name}]",
            f"Columns: {' | '.join(str(c) for c in df.columns)}",
            f"Rows: {len(df)}",
            "",
        ]
        for _, row in df.iterrows():
            lines.append(" | ".join(str(v) for v in row.values))

        log.info(f"  Extracted {len(df)} rows from CSV file")
        return "\n".join(lines), metadata

    except Exception as e:
        log.error(f"  [ERROR] Failed to extract CSV file: {e}")
        return f"[Could not read {path.name}: {e}]", metadata


def _extract_txt(path: Path, metadata: dict) -> tuple[str, dict]:
    """Read a plain text or markdown file directly."""
    try:
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = path.read_text(encoding="latin-1")

        metadata["page_count"] = 1
        log.info(f"  Extracted {len(text):,} characters from text file")
        return text, metadata

    except Exception as e:
        log.error(f"  [ERROR] Failed to read text file: {e}")
        return f"[Could not read {path.name}: {e}]", metadata
