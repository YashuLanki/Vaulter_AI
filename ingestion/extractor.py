"""
ingestion/extractor.py
----------------------
Handles all file text extraction for the Vaulterup ingestion pipeline.

Supported file types:
  .pdf  — pdfplumber for text-based PDFs, Tesseract OCR for scanned/image PDFs
  .xlsx — openpyxl for modern Excel files
  .xls  — xlrd for older Excel files
  .csv  — pandas for comma-separated data
  .txt  — plain text read

All file types are extracted into plain text, then passed through the
same chunker and embedder as before.
"""

import itertools
import logging
from datetime import datetime
from pathlib import Path

import pdfplumber
import pytesseract
from pdf2image import convert_from_path

from config import TESSERACT_PATH, POPPLER_PATH

# Point pytesseract to the correct Tesseract executable
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

log = logging.getLogger("vaulter.extractor")

# ─── Supported File Types ─────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv", ".txt"}


def is_supported(path: Path) -> bool:
    """Return True if this file type is supported by the extractor."""
    return path.suffix.lower() in SUPPORTED_EXTENSIONS


# ─── Main Extraction Entry Point ──────────────────────────────────────────────

def extract(path: Path) -> tuple[str, dict]:
    """
    Main extraction entry point.
    Routes each file to the correct extractor based on its extension.
    Returns (full_text, metadata_dict).
    """
    ext = path.suffix.lower()

    metadata = {
        "filename":    path.name,
        "file_type":   ext,
        "ingested_at": datetime.now().isoformat(),
        "page_count":  0,
        "has_tables":  False,
        "ocr_used":    False,
    }

    if ext == ".pdf":
        return _extract_pdf(path, metadata)
    elif ext in (".xlsx", ".xls"):
        return _extract_excel(path, metadata)
    elif ext == ".csv":
        return _extract_csv(path, metadata)
    elif ext == ".txt":
        return _extract_txt(path, metadata)
    else:
        log.warning(f"  [WARN] Unsupported file type: {ext}")
        return "", metadata


# ─── PDF Extraction ───────────────────────────────────────────────────────────

def _extract_pdf(path: Path, metadata: dict) -> tuple[str, dict]:
    """
    Extract each page with pdfplumber. Any individual page that yields no
    text layer (e.g. a scanned image page mixed into an otherwise
    digital PDF) falls back to Tesseract OCR for that page only, so a
    mostly-digital PDF with a few scanned pages doesn't silently drop
    those pages -- only a whole-document OCR fallback would have caught
    an all-scanned PDF, missing the mixed case.
    """
    full_text = []
    ocr_page_images = None  # lazily rendered only if some page needs it

    with pdfplumber.open(path) as pdf:
        metadata["page_count"] = len(pdf.pages)

        if pdf.metadata:
            metadata["pdf_title"]  = pdf.metadata.get("Title", "") or ""
            metadata["pdf_author"] = pdf.metadata.get("Author", "") or ""

        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()

            if text and text.strip():
                full_text.append(f"[Page {page_num}]\n{text.strip()}")
            else:
                log.info(f"  Page {page_num} has no text layer — running OCR...")
                if ocr_page_images is None:
                    ocr_page_images = convert_from_path(
                        str(path), dpi=300, poppler_path=POPPLER_PATH
                    )
                    metadata["ocr_used"] = True

                if page_num <= len(ocr_page_images):
                    ocr_text = pytesseract.image_to_string(
                        ocr_page_images[page_num - 1], lang="eng"
                    )
                    if ocr_text.strip():
                        full_text.append(f"[Page {page_num} - OCR]\n{ocr_text.strip()}")

            tables = page.extract_tables()
            if tables:
                metadata["has_tables"] = True
                for table in tables:
                    table_text = _table_to_text(table, page_num)
                    if table_text:
                        full_text.append(table_text)

    return "\n\n".join(full_text), metadata


def _table_to_text(table: list, page_num: int) -> str:
    """Convert a pdfplumber table (list of lists) into readable plain text."""
    if not table:
        return ""
    lines = [f"[Table on Page {page_num}]"]
    for row in table:
        cleaned = [str(cell).strip() if cell else "" for cell in row]
        lines.append(" | ".join(cleaned))
    return "\n".join(lines)


# ─── Excel Extraction ─────────────────────────────────────────────────────────

def _extract_excel(path: Path, metadata: dict) -> tuple[str, dict]:
    """
    Extract all sheets and cells from an Excel file (.xlsx or .xls).
    Each sheet is converted to readable plain text with rows and columns.
    """
    import openpyxl

    full_text = []
    metadata["has_tables"] = True

    try:
        wb_formulas = None
        if path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(path, data_only=True)
            # data_only=True returns None for any formula cell that was
            # never recalculated/saved by Excel (e.g. a workbook generated
            # programmatically and never opened in Excel) -- a row made up
            # entirely of such cells looks completely empty and would be
            # silently skipped below, even though it has real (just
            # uncached) data. Load a second, formula-preserving copy so we
            # can tell "genuinely blank row" apart from "all-uncalculated-
            # formula row" and fall back to showing the formula text
            # itself rather than losing the row entirely.
            wb_formulas = openpyxl.load_workbook(path, data_only=False)
        else:
            # .xls — convert via openpyxl after reading with xlrd
            import xlrd
            xls_wb = xlrd.open_workbook(str(path))
            wb = _convert_xls_to_openpyxl(xls_wb)

        sheet_count = len(wb.sheetnames)
        metadata["page_count"] = sheet_count

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines = [f"[Sheet: {sheet_name}]"]

            formula_rows = wb_formulas[sheet_name].iter_rows(values_only=True) if wb_formulas else iter(())
            for row, formula_row in itertools.zip_longest(ws.iter_rows(values_only=True), formula_rows, fillvalue=()):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    if any(isinstance(c, str) and c.startswith("=") for c in formula_row):
                        # Not actually empty -- every cell is an
                        # uncalculated formula. Show the formula text
                        # itself since we can't evaluate it ourselves.
                        cleaned = [str(c) if c is not None else "" for c in formula_row]
                        sheet_lines.append(" | ".join(cleaned))
                    continue
                cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
                sheet_lines.append(" | ".join(cleaned))

            if len(sheet_lines) > 1:  # More than just the header
                # Use double newline (\n\n) between rows so the chunker
                # treats each row as its own paragraph and keeps it intact.
                # Single \n made the entire sheet one giant paragraph,
                # causing the chunker to cut rows in half at arbitrary
                # character positions — breaking pipe-separated data extraction.
                full_text.append("\n\n".join(sheet_lines))

        # Set a sentinel page_count that triggers the 8000-char chunk tier
        # in config.CHUNK_TIERS. Excel rows average 1,600 chars (CoStar
        # exports run to ~3,000 chars/row). The default 500-char chunk size
        # fragments every row — 8,000 keeps each row intact as one chunk.
        metadata["page_count"] = 9999

        log.info(f"  Extracted {sheet_count} sheet(s) from Excel file")

    except Exception as e:
        log.error(f"  [ERROR] Failed to extract Excel file: {e}")
        return "", metadata

    return "\n\n".join(full_text), metadata


def _convert_xls_to_openpyxl(xls_wb):
    """Convert an xlrd workbook to openpyxl format for uniform processing."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_idx in range(xls_wb.nsheets):
        xls_sheet = xls_wb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=xls_sheet.name)
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                ws.cell(row=row + 1, column=col + 1, value=xls_sheet.cell_value(row, col))

    return wb


# ─── CSV Extraction ───────────────────────────────────────────────────────────

def _extract_csv(path: Path, metadata: dict) -> tuple[str, dict]:
    """
    Extract data from a CSV file using pandas.
    Converts rows and columns into readable plain text.
    """
    import pandas as pd

    metadata["has_tables"] = True

    try:
        # Try UTF-8 first, fall back to latin-1 for older exports
        try:
            df = pd.read_csv(path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, encoding="latin-1")

        metadata["page_count"] = 1

        lines = [f"[CSV File: {path.name}]"]
        lines.append(f"Columns: {' | '.join(str(c) for c in df.columns)}")
        lines.append(f"Rows: {len(df)}")
        lines.append("")

        # Include all rows as readable text
        for _, row in df.iterrows():
            row_text = " | ".join(str(v) for v in row.values)
            lines.append(row_text)

        log.info(f"  Extracted {len(df)} rows from CSV file")
        return "\n".join(lines), metadata

    except Exception as e:
        log.error(f"  [ERROR] Failed to extract CSV file: {e}")
        return "", metadata


# ─── Plain Text Extraction ────────────────────────────────────────────────────

def _extract_txt(path: Path, metadata: dict) -> tuple[str, dict]:
    """Read a plain text file directly."""
    try:
        try:
            text = path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = path.read_text(encoding="latin-1")

        metadata["page_count"] = 1
        log.info(f"  Extracted {len(text):,} characters from text file")
        return text, metadata

    except Exception as e:
        log.error(f"  [ERROR] Failed to read text file: {e}")
        return "", metadata