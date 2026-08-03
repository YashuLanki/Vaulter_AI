"""
portfolio.py
------------
The firm's property list, read from the Smartsheet "Project Master" export in
`data/project_master/`.

Extracted from the old `pipeline/property_scraper.py`, which mixed two
unrelated jobs in one 1,000-line module: reading the Project Master (this
file) and scraping news headlines per property (deleted -- see
docs/REBUILD_PLAN.md on why national brokerage news is the wrong signal for
raw land).

What was dropped in the move: the PDF and OCR parsing paths, roughly 250 lines
including rendered-pixel strikethrough detection. Those existed because the
Project Master used to arrive as a scanned PDF export. It is a clean CSV now.
If a PDF export ever comes back, re-export it as CSV or .xlsx instead of
reviving that code.

Known limitation, inherited not introduced: a CSV export cannot represent
strikethrough, which is how the Smartsheet marks a deal sold. So a CSV
Project Master yields every row as active and an empty sold list. Only the
.xlsx path can tell the difference (`cell.font.strike`). Export .xlsx if the
sold/active split matters.
"""

import csv
import logging
from pathlib import Path

from config import DATA_DIR, SMARTSHEET_PORTFOLIO_DIR

log = logging.getLogger("vaulter.portfolio")

PROJECT_MASTER_DIR = DATA_DIR / "project_master"

# Lives in the same folder but is NOT a Project Master -- it's the geocoded
# coordinates table owned by pipeline/property_coordinates.py. Excluded by
# name here because it isn't hypothetical: once that file was added,
# find_project_file() picked it (iterdir order), load_properties() found no
# "Project Name" column, and the portfolio MCP tools failed with
# "Could not extract any properties from property_coordinates.csv".
COORDS_FILENAME = "property_coordinates.csv"

# Also lives alongside the Project Master and is also not one. Excluded by name
# for the same reason as COORDS_FILENAME above: without this, a Project Master
# whose filename doesn't happen to contain "project" and "master" could lose the
# tie-break below to builtin_properties.json and be silently ignored.
BUILTIN_PROPERTIES_FILENAME = "builtin_properties.json"


def _portfolio_dirs() -> list:
    """
    Everywhere the firm's portfolio data can live, in priority order.

    Local first, so a machine that already has a file keeps behaving exactly as
    it did and a file someone deliberately dropped on their own machine always
    beats the team copy. The shared "Smartsheet Portfolio" folder second --
    that's what makes a fresh install useful at all, since the handoff package
    ships no firm data by design.
    """
    dirs = [PROJECT_MASTER_DIR]
    if SMARTSHEET_PORTFOLIO_DIR not in dirs:
        dirs.append(SMARTSHEET_PORTFOLIO_DIR)
    return dirs


# ─── Known properties ─────────────────────────────────────────────────────────
# The firm's real property list used to be a hardcoded literal here. Moved to
# data/project_master/builtin_properties.json (gitignored) on 2026-07-29: this
# repo is deliberately public, and the actual portfolio -- 52 real property
# names, cities, and deal stages -- is confidential business data that has no
# business in a public file. The JSON carries the same list of
# {name, city, state, category} dicts this literal used to hold.
#
# Two jobs, unchanged:
#   * CITY_OVERRIDES -- the CSV export has no city column, only state, so this
#     is the only place a property's city comes from.
#   * Filling in details for sold properties, which the .xlsx parser reports
#     by name only.
#
# Without the JSON (e.g. a fresh public clone), everything still works --
# cities just fall back to the state name, exactly what already happened for
# any property that wasn't in the hardcoded list.

BUILTIN_PROPERTIES_FILE = PROJECT_MASTER_DIR / BUILTIN_PROPERTIES_FILENAME


def _find_portfolio_file(filename: str):
    """First match across _portfolio_dirs(), or None."""
    for directory in _portfolio_dirs():
        candidate = directory / filename
        try:
            if candidate.exists():
                return candidate
        except OSError:
            continue  # an unreachable shared folder is not an error here
    return None


def _load_builtin_properties() -> list[dict]:
    found = _find_portfolio_file(BUILTIN_PROPERTIES_FILENAME)
    if found is None:
        log.info(f"No {BUILTIN_PROPERTIES_FILENAME} -- city overrides and sold-property "
                 f"details unavailable (cities will show as the state name)")
        return []
    try:
        import json
        return json.loads(found.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning(f"Could not read {found}: {e}")
        return []


BUILTIN_PROPERTIES = _load_builtin_properties()
CITY_OVERRIDES = {p["name"]: p["city"] for p in BUILTIN_PROPERTIES}


# ─── Finding and parsing the export ───────────────────────────────────────────

def find_project_file() -> Path | None:
    """
    The Project Master export -- the local copy if there is one, otherwise the
    team's shared "Smartsheet Portfolio" folder. None if neither has one.
    """
    for directory in _portfolio_dirs():
        try:
            directory.mkdir(parents=True, exist_ok=True)
            files = [
                f for f in directory.iterdir()
                if f.is_file() and not f.name.startswith(".")
                and f.name not in (COORDS_FILENAME, BUILTIN_PROPERTIES_FILENAME)
            ]
        except OSError as e:
            # An unreachable shared folder (OneDrive signed out, offline) is a
            # reason to try the next location, not to fail the whole lookup.
            log.warning(f"Could not read {directory}: {e}")
            continue
        if not files:
            continue
        if len(files) > 1:
            # Prefer a file that actually looks like the Project Master rather
            # than whatever iterdir happens to return first.
            named = [f for f in files if "project" in f.stem.lower() and "master" in f.stem.lower()]
            if named:
                files = named
            else:
                log.warning(f"Multiple files in {directory} — using: {files[0].name}")
        return files[0]
    return None


def _missing_file_error() -> FileNotFoundError:
    return FileNotFoundError(
        "\nNo Project Master file found. Looked in both:\n"
        f"  {SMARTSHEET_PORTFOLIO_DIR}   (shared with the team)\n"
        f"  {PROJECT_MASTER_DIR}   (this machine only)\n\n"
        "Usually this means nobody has published the export to the shared\n"
        "folder yet. Export the Vaulter Project Master from Smartsheet and\n"
        "drop it into the shared folder above -- that makes it work for the\n"
        "whole team at once, not just this machine.\n"
        "Supported formats: CSV, Excel (.xlsx). Export .xlsx if you need\n"
        "sold/struck-through deals separated from active ones.\n"
    )


def parse_csv(path: Path) -> list[dict]:
    properties = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [h.strip() if h else h for h in reader.fieldnames]
        for row in reader:
            row      = {k: (v.strip() if v else "") for k, v in row.items()}
            name     = row.get("Project Name", "")
            category = row.get("Project Category", "")
            state    = row.get("State", "")
            if not name or not category or name.lower() == "template":
                continue
            properties.append({
                "name":     name,
                "city":     CITY_OVERRIDES.get(name, state),
                "state":    state,
                "category": category,
            })
    return properties


def parse_excel(path: Path) -> tuple[list[dict], set]:
    """
    Returns (active_properties, sold_names).

    Sold deals are marked in the Smartsheet export by strikethrough on the
    row. An .xlsx carries that as real font formatting openpyxl can read via
    cell.font.strike, so iter_rows() here is deliberately NOT called with
    values_only=True -- that would give only the text and silently lose the
    strikethrough signal, which is exactly what made the CSV/Excel Project
    Master never filter out sold properties at all.
    """
    import openpyxl

    wb      = openpyxl.load_workbook(path, data_only=True)
    ws      = wb.active
    rows    = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], set()
    headers = [str(c).strip() if c else "" for c in rows[0]]
    col     = {h: i for i, h in enumerate(headers)}
    name_i  = col.get("Project Name",     0)
    cat_i   = col.get("Project Category", 2)
    state_i = col.get("State",            3)

    properties = []
    sold_names = set()
    for excel_row in ws.iter_rows(min_row=2):
        row = [c.value for c in excel_row]
        if name_i >= len(row):
            continue
        name     = str(row[name_i]).strip()  if row[name_i]  else ""
        category = str(row[cat_i]).strip()   if cat_i   < len(row) and row[cat_i]   else ""
        state    = str(row[state_i]).strip() if state_i < len(row) and row[state_i] else ""
        if not name or not category or name.lower() == "template":
            continue

        name_cell = excel_row[name_i]
        if name_cell.font and name_cell.font.strike:
            sold_names.add(name)
            continue

        properties.append({
            "name":     name,
            "city":     CITY_OVERRIDES.get(name, state),
            "state":    state,
            "category": category,
        })
    return properties, sold_names


def parse_text(path: Path) -> list[dict]:
    content = path.read_text(encoding="utf-8", errors="replace")
    return [p for p in BUILTIN_PROPERTIES if p["name"] in content]


def _parse(file: Path) -> tuple[list[dict], set]:
    """Dispatch on extension. Returns (properties, sold_names)."""
    ext = file.suffix.lower()
    if ext == ".csv":
        return parse_csv(file), set()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return parse_excel(file)
    return parse_text(file), set()


# Cached against the export's mtime -- these are called repeatedly within a
# single request (once per property, once per lookup). A cache hit costs one
# stat(); a fresh Smartsheet export dropped in changes the mtime and
# invalidates it automatically.
_cache = {"mtime": None, "active": None, "sold": None, "source": None}


def _load(file: Path) -> None:
    mtime = file.stat().st_mtime
    if _cache["active"] is not None and _cache["mtime"] == mtime:
        return

    log.info(f"Loading properties from: {file.name}")
    try:
        active, sold_names = _parse(file)
    except Exception as e:
        raise ValueError(f"Failed to parse {file.name}: {e}") from e

    if not active:
        raise ValueError(f"Could not extract any properties from {file.name}")

    builtin_map = {p["name"]: p for p in BUILTIN_PROPERTIES}
    sold = [
        {**builtin_map[name], "status": "sold"} if name in builtin_map
        else {"name": name, "city": CITY_OVERRIDES.get(name, "unknown"),
              "state": "unknown", "category": "unknown", "status": "sold"}
        for name in sorted(sold_names)
    ]

    if sold_names:
        log.info(f"{len(sold_names)} struck-through properties excluded: {', '.join(sorted(sold_names))}")
    log.info(f"Loaded {len(active)} active, {len(sold)} sold from {file.name}")

    _cache.update(mtime=mtime, active=active, sold=sold, source=file.name)


def load_properties() -> tuple[list[dict], str]:
    """
    Active properties plus the name of the file they came from.
    Raises FileNotFoundError if no export is present.
    """
    file = find_project_file()
    if not file:
        raise _missing_file_error()
    _load(file)
    return _cache["active"], _cache["source"]


def load_all_properties() -> tuple[list[dict], list[dict]]:
    """
    (active, sold). Sold deals carry status="sold".

    Sold is always empty for a CSV export -- see this module's header.
    """
    file = find_project_file()
    if not file:
        raise _missing_file_error()
    _load(file)
    return _cache["active"], _cache["sold"]
